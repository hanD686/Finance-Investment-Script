from __future__ import annotations

import argparse
import json
import re
import shutil
import socket
import subprocess
import sys
from datetime import date, datetime, timedelta
from html import escape
from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
DEFAULT_CONFIG = {
    "page_size": 50,
    "output_dir": "output",
    "fund_start_date": "2024-01-01",
    "fund_end_date": "",
    "source_pages": {
        "quote": "https://quote.eastmoney.com/center/",
        "data": "https://data.eastmoney.com/center/",
        "fund": "https://fund.eastmoney.com/?spm=100015.lw.4",
    },
    "market_fs": "m:0+t:6,m:0+t:80,m:1+t:2,m:1+t:23",
}


QUOTE_FIELDS = {
    "f12": "代码",
    "f14": "名称",
    "f2": "最新价",
    "f3": "涨跌幅%",
    "f4": "涨跌额",
    "f5": "成交量(手)",
    "f6": "成交额",
    "f15": "最高",
    "f16": "最低",
    "f17": "今开",
    "f18": "昨收",
    "f8": "换手率%",
    "f9": "市盈率",
    "f23": "市净率",
    "f20": "总市值",
    "f21": "流通市值",
}

MONEY_FIELDS = {
    "f12": "代码",
    "f14": "名称",
    "f2": "最新价",
    "f3": "涨跌幅%",
    "f62": "主力净流入",
    "f184": "主力净占比%",
    "f66": "超大单净流入",
    "f69": "超大单净占比%",
    "f72": "大单净流入",
    "f75": "大单净占比%",
    "f78": "中单净流入",
    "f81": "中单净占比%",
    "f84": "小单净流入",
    "f87": "小单净占比%",
}

FUND_COLUMNS = [
    "代码",
    "名称",
    "拼音",
    "净值日期",
    "单位净值",
    "累计净值",
    "日增长率%",
    "近1周%",
    "近1月%",
    "近3月%",
    "近6月%",
    "近1年%",
    "近2年%",
    "近3年%",
    "今年来%",
    "成立来%",
    "成立日期",
    "自定义区间%",
    "手续费",
]


class EastmoneyClient:
    def __init__(self) -> None:
        self.session = requests.Session()
        self.session.trust_env = False
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126 Safari/537.36"
                ),
                "Accept": "application/json,text/javascript,*/*;q=0.1",
                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            }
        )

    def get_json(self, url: str, params: dict[str, Any], referer: str) -> dict[str, Any]:
        try:
            response = self.session.get(url, params=params, headers={"Referer": referer}, timeout=30)
            response.raise_for_status()
            return response.json()
        except requests.RequestException:
            return json.loads(self.curl_get(url, params, referer))

    def get_text(self, url: str, params: dict[str, Any], referer: str) -> str:
        try:
            response = self.session.get(url, params=params, headers={"Referer": referer}, timeout=30)
            response.raise_for_status()
            response.encoding = response.apparent_encoding or "utf-8"
            return response.text
        except requests.RequestException:
            return self.curl_get(url, params, referer)

    def curl_get(self, url: str, params: dict[str, Any], referer: str) -> str:
        curl = shutil.which("curl.exe") or shutil.which("curl")
        if not curl:
            raise RuntimeError("requests 请求失败，并且系统里找不到 curl/curl.exe 作为回退工具。")
        full_url = f"{url}?{urlencode(params)}"
        result = subprocess.run(
            [
                curl,
                "-sS",
                "-L",
                "--max-time",
                "30",
                "-A",
                self.session.headers["User-Agent"],
                "-e",
                referer,
                full_url,
            ],
            check=True,
            capture_output=True,
            encoding="utf-8",
            errors="replace",
        )
        return result.stdout


def load_config(config_path: Path) -> dict[str, Any]:
    config = json.loads(json.dumps(DEFAULT_CONFIG))
    if config_path.exists():
        with config_path.open("r", encoding="utf-8") as file:
            user_config = json.load(file)
        deep_update(config, user_config)
    return config


def deep_update(target: dict[str, Any], source: dict[str, Any]) -> None:
    for key, value in source.items():
        if isinstance(value, dict) and isinstance(target.get(key), dict):
            deep_update(target[key], value)
        else:
            target[key] = value


def fetch_quote_rows(client: EastmoneyClient, config: dict[str, Any]) -> list[dict[str, Any]]:
    fields = ",".join(QUOTE_FIELDS)
    data = client.get_json(
        "https://push2.eastmoney.com/api/qt/clist/get",
        {
            "pn": 1,
            "pz": config["page_size"],
            "po": 1,
            "np": 1,
            "ut": "bd1d9ddb04089700cf9c27f6f7426281",
            "fltt": 2,
            "invt": 2,
            "fid": "f3",
            "fs": config["market_fs"],
            "fields": fields,
        },
        config["source_pages"]["quote"],
    )
    return map_rows(data, QUOTE_FIELDS)


def fetch_money_rows(client: EastmoneyClient, config: dict[str, Any]) -> list[dict[str, Any]]:
    fields = ",".join(MONEY_FIELDS)
    data = client.get_json(
        "https://push2.eastmoney.com/api/qt/clist/get",
        {
            "pn": 1,
            "pz": config["page_size"],
            "po": 1,
            "np": 1,
            "ut": "bd1d9ddb04089700cf9c27f6f7426281",
            "fltt": 2,
            "invt": 2,
            "fid": "f62",
            "fs": config["market_fs"],
            "fields": fields,
        },
        config["source_pages"]["data"],
    )
    return map_rows(data, MONEY_FIELDS)


def fetch_fund_rows(client: EastmoneyClient, config: dict[str, Any]) -> list[dict[str, Any]]:
    end_date = parse_config_date(config.get("fund_end_date")) or date.today()
    start_date = parse_config_date(config.get("fund_start_date")) or (end_date - timedelta(days=365))

    last_text = ""
    for offset in range(0, 8):
        candidate_end = end_date - timedelta(days=offset)
        text = client.get_text(
            "https://fund.eastmoney.com/data/rankhandler.aspx",
            {
                "op": "ph",
                "dt": "kf",
                "ft": "all",
                "rs": "",
                "gs": "0",
                "sc": "zzf",
                "st": "desc",
                "sd": start_date.isoformat(),
                "ed": candidate_end.isoformat(),
                "qdii": "",
                "tabSubtype": ",,,,,",
                "pi": 1,
                "pn": config["page_size"],
                "dx": 1,
                "v": f"{datetime.now().timestamp():.6f}",
            },
            "https://fund.eastmoney.com/data/fundranking.html",
        )
        last_text = text
        rows = parse_fund_rank(text)
        if rows:
            return rows

    raise RuntimeError(f"基金排行接口连续多天无数据，最后响应前 200 字：{last_text[:200]}")


def map_rows(payload: dict[str, Any], field_map: dict[str, str]) -> list[dict[str, Any]]:
    rows = payload.get("data", {}).get("diff") or []
    return [{label: clean_value(raw.get(field)) for field, label in field_map.items()} for raw in rows]


def parse_fund_rank(text: str) -> list[dict[str, Any]]:
    match = re.search(r"datas:\[(?P<datas>.*?)\],allRecords", text, flags=re.S)
    if not match:
        return []
    data_block = match.group("datas").strip()
    if not data_block:
        return []

    raw_rows = json.loads(f"[{data_block}]")
    parsed: list[dict[str, Any]] = []
    for raw in raw_rows:
        values = raw.split(",")
        row = {}
        for index, column in enumerate(FUND_COLUMNS):
            source_index = 17 if column == "自定义区间%" else index
            if column == "手续费":
                source_index = 19
            row[column] = clean_value(values[source_index] if source_index < len(values) else "")
        parsed.append(row)
    return parsed


def parse_config_date(value: Any) -> date | None:
    if not value:
        return None
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value), "%Y-%m-%d").date()


def clean_value(value: Any) -> Any:
    if value in (None, "-", ""):
        return ""
    return value


def build_snapshot(config: dict[str, Any]) -> dict[str, Any]:
    client = EastmoneyClient()
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return {
        "generated_at": generated_at,
        "source_pages": config["source_pages"],
        "tables": {
            "A股行情涨幅榜": fetch_quote_rows(client, config),
            "数据中心-主力资金净流入": fetch_money_rows(client, config),
            "基金排行": fetch_fund_rows(client, config),
        },
    }


def write_excel(snapshot: dict[str, Any], output_dir: Path) -> Path:
    daily_dir = output_dir / "daily"
    daily_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    workbook_path = daily_dir / f"eastmoney_daily_{stamp}.xlsx"

    workbook = Workbook()
    workbook.remove(workbook.active)

    metadata = workbook.create_sheet("说明")
    metadata.append(["生成时间", snapshot["generated_at"]])
    metadata.append(["数据来源", "东方财富网页及其公开页面数据接口"])
    for name, url in snapshot["source_pages"].items():
        metadata.append([name, url])
    style_sheet(metadata)

    for sheet_name, rows in snapshot["tables"].items():
        worksheet = workbook.create_sheet(safe_sheet_name(sheet_name))
        write_table(worksheet, rows)

    workbook.save(workbook_path)
    latest_path = output_dir / "latest.xlsx"
    shutil.copyfile(workbook_path, latest_path)
    return workbook_path


def write_table(worksheet: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        worksheet.append(["暂无数据"])
        return

    headers = list(rows[0].keys())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    style_sheet(worksheet)


def style_sheet(worksheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    for index, column_cells in enumerate(worksheet.columns, start=1):
        width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 10), 28)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def write_html(snapshot: dict[str, Any], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    parts = [
        "<!doctype html>",
        '<html lang="zh-CN">',
        "<head>",
        '<meta charset="utf-8">',
        '<meta name="viewport" content="width=device-width, initial-scale=1">',
        "<title>东方财富每日行情</title>",
        "<style>",
        CSS,
        "</style>",
        "</head>",
        "<body>",
        "<main>",
        "<header>",
        "<div>",
        "<p>Eastmoney Daily</p>",
        "<h1>东方财富每日行情</h1>",
        f"<time>更新时间：{escape(snapshot['generated_at'])}</time>",
        "</div>",
        '<a class="download" href="latest.xlsx">下载 Excel</a>',
        "</header>",
        '<section class="sources">',
    ]
    for label, url in snapshot["source_pages"].items():
        parts.append(f'<a href="{escape(url)}" target="_blank" rel="noreferrer">{escape(label)} 来源</a>')
    parts.append("</section>")

    for title, rows in snapshot["tables"].items():
        parts.append(render_table(title, rows))

    parts.extend(["</main>", "</body>", "</html>"])
    html_path = output_dir / "index.html"
    html_path.write_text("\n".join(parts), encoding="utf-8")
    return html_path


def render_table(title: str, rows: list[dict[str, Any]]) -> str:
    if not rows:
        return f"<section><h2>{escape(title)}</h2><p>暂无数据</p></section>"
    headers = list(rows[0].keys())
    html = [f"<section><h2>{escape(title)}</h2>", '<div class="table-wrap"><table><thead><tr>']
    html.extend(f"<th>{escape(header)}</th>" for header in headers)
    html.append("</tr></thead><tbody>")
    for row in rows:
        html.append("<tr>")
        html.extend(f"<td>{escape(str(row.get(header, '')))}</td>" for header in headers)
        html.append("</tr>")
    html.append("</tbody></table></div></section>")
    return "\n".join(html)


def safe_sheet_name(name: str) -> str:
    return re.sub(r"[\[\]*?:/\\]", "-", name)[:31]


def serve(output_dir: Path, host: str, port: int) -> None:
    class Handler(SimpleHTTPRequestHandler):
        def __init__(self, *args: Any, **kwargs: Any) -> None:
            super().__init__(*args, directory=str(output_dir), **kwargs)

    server = ThreadingHTTPServer((host, port), Handler)
    print(f"Serving {output_dir}")
    for url in local_urls(port):
        print(f"Preview: {url}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")


def local_urls(port: int) -> list[str]:
    urls = [f"http://127.0.0.1:{port}/"]
    try:
        hostname = socket.gethostname()
        for _, _, _, _, sockaddr in socket.getaddrinfo(hostname, None, family=socket.AF_INET):
            ip = sockaddr[0]
            if not ip.startswith("127.") and f"http://{ip}:{port}/" not in urls:
                urls.append(f"http://{ip}:{port}/")
    except OSError:
        pass
    return urls


def update(config_path: Path) -> None:
    config = load_config(config_path)
    output_dir = (ROOT / config["output_dir"]).resolve()
    snapshot = build_snapshot(config)
    workbook_path = write_excel(snapshot, output_dir)
    html_path = write_html(snapshot, output_dir)
    print(f"Excel: {workbook_path}")
    print(f"Latest: {output_dir / 'latest.xlsx'}")
    print(f"HTML: {html_path}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="东方财富每日行情采集、Excel 导出和移动端预览")
    parser.add_argument("--config", default="config.json", help="配置文件路径，默认 config.json")
    subparsers = parser.add_subparsers(dest="command")

    subparsers.add_parser("update", help="采集数据并生成 Excel/HTML")

    serve_parser = subparsers.add_parser("serve", help="启动本地网页服务，供手机同网段访问")
    serve_parser.add_argument("--host", default="0.0.0.0")
    serve_parser.add_argument("--port", default=8000, type=int)

    args = parser.parse_args(argv)
    command = args.command or "update"
    config_path = (ROOT / args.config).resolve()
    config = load_config(config_path)

    if command == "update":
        update(config_path)
        return 0
    if command == "serve":
        serve((ROOT / config["output_dir"]).resolve(), args.host, args.port)
        return 0
    parser.error(f"未知命令：{command}")
    return 2


CSS = """
:root {
  color-scheme: light;
  --ink: #18212f;
  --muted: #637083;
  --line: #d7dee8;
  --surface: #ffffff;
  --brand: #0d6a8d;
  --accent: #c74b35;
  --bg: #f3f6f8;
}
* { box-sizing: border-box; }
body {
  margin: 0;
  color: var(--ink);
  background: var(--bg);
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft YaHei", sans-serif;
}
main {
  width: min(1180px, 100%);
  margin: 0 auto;
  padding: 20px;
}
header {
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 16px;
  padding: 18px 0 14px;
  border-bottom: 1px solid var(--line);
}
header p {
  margin: 0 0 4px;
  color: var(--accent);
  font-size: 13px;
  font-weight: 700;
}
h1 {
  margin: 0;
  font-size: clamp(26px, 4vw, 42px);
  line-height: 1.1;
}
time {
  display: block;
  margin-top: 8px;
  color: var(--muted);
  font-size: 14px;
}
.download {
  flex: 0 0 auto;
  padding: 10px 14px;
  border-radius: 6px;
  color: #fff;
  background: var(--brand);
  font-weight: 700;
  text-decoration: none;
}
.sources {
  display: flex;
  flex-wrap: wrap;
  gap: 10px;
  padding: 14px 0 2px;
}
.sources a {
  color: var(--brand);
  font-size: 14px;
}
section {
  margin-top: 22px;
}
h2 {
  margin: 0 0 10px;
  font-size: 20px;
}
.table-wrap {
  overflow-x: auto;
  border: 1px solid var(--line);
  border-radius: 8px;
  background: var(--surface);
}
table {
  width: 100%;
  min-width: 900px;
  border-collapse: collapse;
  font-size: 13px;
}
th, td {
  padding: 9px 10px;
  border-bottom: 1px solid var(--line);
  white-space: nowrap;
  text-align: right;
}
th:first-child, td:first-child,
th:nth-child(2), td:nth-child(2) {
  text-align: left;
}
th {
  position: sticky;
  top: 0;
  z-index: 1;
  color: #fff;
  background: #1f4e78;
  font-weight: 700;
}
tr:nth-child(even) td {
  background: #f8fafc;
}
@media (max-width: 640px) {
  main { padding: 14px; }
  header {
    align-items: flex-start;
    flex-direction: column;
  }
  .download {
    width: 100%;
    text-align: center;
  }
  table {
    min-width: 760px;
    font-size: 12px;
  }
}
"""


if __name__ == "__main__":
    raise SystemExit(main())
